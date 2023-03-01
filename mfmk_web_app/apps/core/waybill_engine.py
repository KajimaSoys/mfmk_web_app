from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NumberFormatDescriptor


def fill_waybill(in_file, data, out_file):
    """Base control settings"""

    wb = load_workbook(in_file, read_only=False, keep_vba=True)
    sheet = wb.get_sheet_by_name('Waybill')
    wb.active = sheet
    ws = wb.active

    fill_positions(ws, data)

    wb.save(out_file)
    return True




def fill_positions(ws, data):
    """Generating control settings"""

    type_choices = {
        'closet': 'Шкаф',
        'switching': 'Коммутационное оборудование',
        'softstarter': 'Пч и упп',
        'controllers': 'Контроллеры',
        'avr': 'Авр',
        'tires': 'Шины',
        'signalfittings': 'Сигнальная арматура и переключатели',
        'relays': 'Клеммы и реле',
        'consumables': 'Расходные материалы',
    }

    discount = {
        'Корпус навесной металлический 1000х600х300 IP66 RAL7035': '0',
        'Выключатель-разъединитель 3Р 63А ВН-102': '35',
        'Авт. выкл. защ. двиг. 3P 40-63A 30кА ВА-432': '35',
        'Контакт доп. боковой 1НО+1НЗ для ВА-432': '35',
        'Авт. выкл. защ. двиг. 3P 1,6-2,5A 100кА ВА-431': '35',
        'Контакт доп. боковой 1НО+1НЗ для ВА-431': '35',
        'Контактор 9А 220В/230В АС3 АС4 1НО КМ-102': '35',
        'Приставка контактная доп.контакты 2НО+2НЗ лиц. ПК03-02-22': '35',
        'УСТРОЙСТВО ПЛАВНОГО ПУСКА ESQ-GS7-030 30 кВт': '15',
        'Блок питания MDR-60-24, 60 Вт, 2,4 А': '0',
        'EDS-205A Компактный коммутатор 5*10/100BaseTX в метал.корп': '0',
        'ОПЛК Vision V570 (3xAI, 18xDI, 17xDO)': '0',
        'FLC': '0',
        'Реле контроля напряжения, G1-SA': '0',
        'Контактор 65А 220В/230В АС3 АС4 1НО+1НЗ КМ-102': '35',
        'Приставка контактная доп.контакты 2НО+2НЗ лиц. ПК03-02-22': '35',
        'Механизм блокировки для контакторов КМ-103 40-95': '35',
        'Светодиод, LED, ~220VAC': '0',
        'KRI200230L; Реле промежуточное; 2C, 5A, LED, 230B AC': '-10',
        'KRI 2CO; Монтажная колодка на Дин-Рейку, KRI; 2C': '-10',
        'Переключатель с 1-0-2': '0',
        'Авт. выкл. 3Р 6А х-ка C ВА-103 6кА': '35',
        'Шина для N и PE (до 100А)': '0',
        'Модуль ПКЛ': '0',
        'Переключатель с 1-0-2': '0',
        'Сдвоенная кнопка без фиксации "старт/стоп", без индикации ': '0',
        'Светодиод, LED, ~220VAC': '0',
        'KD 4 Концевой стопор пружинный': '0',
        'Клемма 1-уровн.,   2,5мм.кв.,   пружинн.,   СЕРАЯ,   (модель PYK 2,5)': '0',
        'KRI200230L; Реле промежуточное; 2C, 5A, LED, 230B AC': '-10',
        'KRI200024L; Реле промежуточное; 2C, 5A, LED, 24B DC': '-10',
        'KRI 2CO; Монтажная колодка на Дин-Рейку, KRI; 2C': '-10',
        'KRM400230L; Реле промежуточное; 4C, 5A, LED, 230B AC': '-10',
        'KRM400024L, Реле промежуточное,4C,5A,LED,24B DC': '-10',
        'KS 4CO; Монтажная колодка на Дин-Рейку; KMY; 4C': '-10',
        'Реле контроля уровня; настраиваемый диапазон чувствительности 5…150кОм; питание 240В AC; выход 1CO 16А; модульное, ширина 35мм; степень защиты IP20': '40',
        'Провода, стяжки, маркировка, метизы, шильды': '0',
    }

    quantity = {
        'closet': {
            'Корпус навесной металлический 1000х600х300 IP66 RAL7035': '1',
        },
        'switching': {
            'Выключатель-разъединитель 3Р 63А ВН-102': '2',
            'Авт. выкл. защ. двиг. 3P 40-63A 30кА ВА-432': '1',
            'Контакт доп. боковой 1НО+1НЗ для ВА-432': '1',
            'Авт. выкл. защ. двиг. 3P 1,6-2,5A 100кА ВА-431': '1',
            'Контакт доп. боковой 1НО+1НЗ для ВА-431': '1',
            'Контактор 9А 220В/230В АС3 АС4 1НО КМ-102': '1',
            'Приставка контактная доп.контакты 2НО+2НЗ лиц. ПК03-02-22': '1',
        },
        'softstarter': {
            'УСТРОЙСТВО ПЛАВНОГО ПУСКА ESQ-GS7-030 30 кВт': '2',
        },
        'controllers': {
            'Блок питания MDR-60-24, 60 Вт, 2,4 А': '1',
            'EDS-205A Компактный коммутатор 5*10/100BaseTX в метал.корп': '1',
            'ОПЛК Vision V570 (3xAI, 18xDI, 17xDO)': '1',
            'FLC': '1',
        },
        'avr': {
            'Реле контроля напряжения, G1-SA': '2',
            'Контактор 65А 220В/230В АС3 АС4 1НО+1НЗ КМ-102': '2',
            'Приставка контактная доп.контакты 2НО+2НЗ лиц. ПК03-02-22': '2',
            'Механизм блокировки для контакторов КМ-103 40-95': '1',
            'Светодиод, LED, ~220VAC': '3',
            'KRI200230L; Реле промежуточное; 2C, 5A, LED, 230B AC': '3',
            'KRI 2CO; Монтажная колодка на Дин-Рейку, KRI; 2C': '3',
            'Переключатель с 1-0-2': '1',
            'Авт. выкл. 3Р 6А х-ка C ВА-103 6кА': '2',
        },
        'tires': {
            'Шина для N и PE (до 100А)': '1',
        },
        'signalfittings': {
            'Модуль ПКЛ': '2',
            'Переключатель с 1-0-2': '3',
            'Сдвоенная кнопка без фиксации "старт/стоп", без индикации ': '3',
            'Светодиод, LED, ~220VAC': '12',
        },
        'relays': {
            'KD 4 Концевой стопор пружинный': '10',
            'Клемма 1-уровн.,   2,5мм.кв.,   пружинн.,   СЕРАЯ,   (модель PYK 2,5)': '50',
            'KRI200230L; Реле промежуточное; 2C, 5A, LED, 230B AC': '3',
            'KRI200024L; Реле промежуточное; 2C, 5A, LED, 24B DC': '15',
            'KRI 2CO; Монтажная колодка на Дин-Рейку, KRI; 2C': '18',
            'KRM400230L; Реле промежуточное; 4C, 5A, LED, 230B AC': '1',
            'KRM400024L, Реле промежуточное,4C,5A,LED,24B DC': '5',
            'KS 4CO; Монтажная колодка на Дин-Рейку; KMY; 4C': '6',
            'Реле контроля уровня; настраиваемый диапазон чувствительности 5…150кОм; питание 240В AC; выход 1CO 16А; модульное, ширина 35мм; степень защиты IP20': '1',
        },
        'consumables': {
            'Провода, стяжки, маркировка, метизы, шильды': '1',
        },
    }

    blue_fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')

    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=medium, right=medium, bottom=thin)


    # Start from here
    row = 3
    i = 1

    for key in data:
        ws[f'B{row}'] = type_choices[key].upper()
        ws[f'B{row}'].font = Font(name='Arial', size=10, b=True)
        for rows in ws.iter_rows(min_row=row, max_row=row, min_col=1, max_col=9):
            for cell in rows:
                cell.fill = blue_fill
        row += 1

        for item in data[key]:
            # number
            ws[f'A{row}'] = i
            ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
            # name
            ws[f'B{row}'] = item.name
            # vendor_code
            ws[f'C{row}'] = item.vendor_code
            ws[f'C{row}'].alignment = Alignment(horizontal="center", vertical="center")
            # manufacturer
            ws[f'D{row}'] = item.manufacturer
            ws[f'D{row}'].alignment = Alignment(horizontal="center", vertical="center")
            # price
            ws[f'E{row}'] = item.price
            # discount
            ws[f'F{row}'] = discount[item.name]
            ws[f'F{row}'].alignment = Alignment(horizontal="center", vertical="center")
            # quantity
            ws[f'G{row}'] = quantity[key][item.name]
            ws[f'G{row}'].alignment = Alignment(horizontal="center", vertical="center")
            # total
            if item.currency == 'rub':
                ws[f'E{row}'].number_format = '# ##0"p."'
                ws[f'H{row}'] = f'=(E{row}-E{row}*F{row}/100)*G{row}'
            elif item.currency == 'usd':
                ws[f'E{row}'].number_format = '#,##0.0" USD"'
                ws[f'H{row}'] = f'=(E{row}-E{row}*F{row}/100)*G{row}*$K$3'
            elif item.currency == 'eur':
                ws[f'E{row}'].number_format = '#,##0.0" EUR"'
                ws[f'H{row}'] = f'=(E{row}-E{row}*F{row}/100)*G{row}*$K$2'
            ws[f'H{row}'].number_format = '# ##0"p."'
            # correction
            ws[f'I{row}'] = f'=H{row}*$K$5'

            i += 1
            row += 1

    set_border(ws, f'A3:I{row-1}')

    # Ends here



def generate_waybill(instance, data):
    """This method generating xlsm document"""

    in_file = 'media/waybill_templates/template.xlsm'
    out_file = f'media/questionnaire/id_{instance.id}/waybill_{instance.id}.xlsm'

    try:
        fill_waybill(in_file, data, out_file)
    except Exception as E:
        print(f"An error occurred: {E}")
        return False
    else:
        return True